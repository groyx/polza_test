"""
Сборка итогового файла для сдачи: один XLSX со всеми листами.

Сдавать нужно Google Таблицей, а XLSX она открывает без потерь — поэтому
собираем книгу локально и заливаем одним файлом, а не копипастим листы
руками.

Форматирование здесь не украшательство. Проверяющий откроет таблицу на
минуту, и от того, видно ли ему сразу структуру, зависит оценка: шапка
закреплена, колонки по ширине содержимого, статусы почты и проблемные
строки подсвечены, длинный текст переносится.

Запуск:
    python src/export_xlsx.py --out data/report.xlsx
"""

from __future__ import annotations

import argparse
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA = ROOT / "data"

HEADER_FILL = PatternFill("solid", fgColor="1F3864")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)

# Подсветка статусов. Красный только там, где строку реально нельзя брать
# в работу, — иначе цвет перестаёт что-либо значить.
FILLS = {
    "valid": PatternFill("solid", fgColor="D8EFD8"),
    "risky": PatternFill("solid", fgColor="FFF3CD"),
    "invalid": PatternFill("solid", fgColor="F8D7DA"),
    "ok": PatternFill("solid", fgColor="D8EFD8"),
    "перепутан сайт": PatternFill("solid", fgColor="F8D7DA"),
    "требует проверки": PatternFill("solid", fgColor="FFF3CD"),
    "не проверено": PatternFill("solid", fgColor="E2E3E5"),
    "нет данных": PatternFill("solid", fgColor="E2E3E5"),
}

# Колонки, которые почти всегда длинные: им даём перенос и фиксированную
# ширину, иначе одна ячейка растягивает лист на два экрана.
WIDE = {
    "Персонализация", "Цитата-подтверждение", "Найденные проблемы",
    "Обоснование", "Примечание", "Заголовок сайта",
    "Как сайт называет себя", "Сигнал (вакансия)",
}


def add_sheet(wb: Workbook, title: str, path: Path, note: str = "") -> bool:
    """Добавляет лист из CSV. Возвращает False, если файла нет."""
    if not path.is_file():
        print(f"  [пропуск] {title}: нет файла {path.name}")
        return False

    with path.open(encoding="utf-8-sig", newline="") as f:
        rows = list(csv.reader(f))
    if not rows:
        print(f"  [пропуск] {title}: файл пуст")
        return False

    ws = wb.create_sheet(title[:31])
    start = 1

    # Пояснение над таблицей: проверяющий видит, что за лист, не спрашивая.
    if note:
        ws.cell(row=1, column=1, value=note).font = Font(italic=True, size=10)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(rows[0])))
        start = 3

    header = rows[0]
    for c, name in enumerate(header, 1):
        cell = ws.cell(row=start, column=c, value=name)
        cell.fill, cell.font = HEADER_FILL, HEADER_FONT
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    status_cols = [
        i for i, h in enumerate(header)
        if h.strip() in ("Статус почты", "Статус", "Способ")
    ]

    for r, row in enumerate(rows[1:], start + 1):
        for c, value in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=value)
            if header[c - 1] in WIDE:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        for i in status_cols:
            if i < len(row):
                fill = FILLS.get(row[i].strip())
                if fill:
                    ws.cell(row=r, column=i + 1).fill = fill

    # Ширина по фактическому содержимому, но с потолком.
    for c, name in enumerate(header, 1):
        if name in WIDE:
            width = 46
        else:
            longest = max(
                (len(str(row[c - 1])) for row in rows if c - 1 < len(row)),
                default=10,
            )
            width = min(max(longest + 2, 12), 40)
        ws.column_dimensions[get_column_letter(c)].width = width

    ws.freeze_panes = ws.cell(row=start + 1, column=1)
    ws.auto_filter.ref = (
        f"A{start}:{get_column_letter(len(header))}{start + len(rows) - 1}"
    )
    print(f"  [ок] {title}: {len(rows) - 1} строк")
    return True


def add_emails_sheet(wb: Workbook, path: Path) -> None:
    """Цепочку писем кладём текстом — так её удобнее читать в таблице."""
    if not path.is_file():
        print(f"  [пропуск] Письма: нет файла {path.name}")
        return
    ws = wb.create_sheet("цепочка писем — письма")
    ws.column_dimensions["A"].width = 118
    for i, line in enumerate(path.read_text(encoding="utf-8").splitlines(), 1):
        cell = ws.cell(row=i, column=1, value=line)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        if line.startswith("## "):
            cell.font = Font(bold=True, size=13)
        elif line.startswith("# "):
            cell.font = Font(bold=True, size=15)
        elif line.startswith("**"):
            cell.font = Font(bold=True)
    print(f"  [ок] Письма: {ws.max_row} строк")


def main() -> None:
    p = argparse.ArgumentParser(description="Сборка итогового XLSX")
    p.add_argument("--out", type=Path, default=DATA / "report.xlsx")
    args = p.parse_args()

    wb = Workbook()
    wb.remove(wb.active)   # пустой лист по умолчанию не нужен

    print("собираю листы:")
    add_sheet(
        wb, "сбор базы — база", DATA / "base_50_personalized.csv",
        "сбор базы и 2. Компании с открытой вакансией менеджера по продажам — "
        "то есть с подтверждённой потребностью в клиентах. "
        "Колонки «Источник» и «Цитата-подтверждение» позволяют проверить любую строку.",
    )
    add_sheet(
        wb, "сбор базы — сырьё", DATA / "base_50.csv",
        "Та же база до персонализации: как её отдал сборщик. "
        "Статус почты: valid — синтаксис и MX в порядке; "
        "risky — ролевой или некорпоративный ящик, брать осознанно.",
    )
    add_sheet(
        wb, "Аудит базы", DATA / "sample_15_audited.csv",
        "аудит базы. Аудит присланной базы. Испорчено 6 строк из 15: "
        "домены и адреса перепутаны между компаниями. "
        "Колонка «Найденные проблемы» — что именно поймал скрипт.",
    )
    add_emails_sheet(wb, ROOT / "emails" / "sequence.md")

    if not wb.sheetnames:
        raise SystemExit("нечего сохранять — сначала запустите пайплайн")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(args.out)
    print(f"\nсохранено: {args.out}")
    print("залейте файл в Google Диск и откройте как Google Таблицу")


if __name__ == "__main__":
    main()
